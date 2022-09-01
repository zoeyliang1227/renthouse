import time

import bs4
import hyperlink
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

timeout = 20


def get_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--headless')                 # 瀏覽器不提供可視化頁面
    options.add_argument('-no-sandbox')               # 以最高權限運行
    options.add_argument('--start-maximized')        # 縮放縮放（全屏窗口）設置元素比較準確
    options.add_argument('--disable-gpu')            # 谷歌文檔說明需要加上這個屬性來規避bug
    options.add_argument('--window-size=1920,1080')  # 設置瀏覽器按鈕（窗口大小）
    options.add_argument('--incognito')               # 啟動無痕

    driver = webdriver.Chrome(chrome_options=options)
    driver.get(
        'https://rent.591.com.tw/?mrtline=100&mrtcoods=4323,4324,4185,4186,4187&searchtype=4&rentprice=9000,12000&other=near_subway&showMore=1&area=8,&multiFloor=2_6,0_1&option=cold,washer,icebox,hotwater,broadband,bed&multiNotice=not_cover,all_sex')

    return driver


def search():
    driver = get_driver()
    wb = openpyxl.Workbook()
    ws = wb.active
    sheet = wb.create_sheet("house", 0)
    # 先填入第一列的欄位名稱
    sheet['A1'] = 'text'
    sheet['B1'] = 'title'
    sheet['C1'] = 'area'
    sheet['D1'] = 'subway'
    sheet['E1'] = 'href'
    sheet['F1'] = 'style'

    # for j in range(300):
    #     driver.execute_script(
    #         'window.scrollTo(0, document.body.scrollHeight);')
    #     time.sleep(1)
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.CLASS_NAME, 'item-title')))
    soup = bs4.BeautifulSoup(driver.page_source, 'html.parser')
    house = soup.find_all('a')
    # print(house)
    # print(len(house))
    # path = 'output.txt'
    # f = open(path, 'w', encoding='UTF-8')
    # f.write(str(house))
    # f.close()

    for i in house:
        # print(i)
        data = i.find('div', class_='rent-item-right')
        # print(data)
        a = i.find('div', class_='item-price-text')
        b = i.find('div', class_='item-title')
        c = i.find('div', class_='item-area')
        d = i.find('div', class_='item-tip subway')
        e = ws.cell(row=1, column=1).value = '=HYPERLINK("{}")'.format(
            i.get('href'))

        f = i.find('ul', class_='item-style')

        if not b:
            pass
        else:
            if not ('林森北' in c.text):
                if not ('雅房' in b.text.strip()):
                    sheet.append([a.text, b.text.strip(), c.text,
                                  d.text.strip(), e, f.text])

    wb.save("house.xlsx")


search()